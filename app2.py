import fitz  # PyMuPDF
from openpyxl import Workbook
import re

# def clean_question(text):
#     text = text.strip()

#     # Remove "Comments:"
#     text = re.sub(r'Comments\s*:\s*', '', text, flags=re.IGNORECASE)

#     # Remove extra spaces
#     text = re.sub(r'\s+', ' ', text)

#     return text.strip()

def clean_question(text):
    text = text.strip()

    # Remove "Comments:"
    text = re.sub(r'Comments\s*:\s*', '', text, flags=re.IGNORECASE)

    # Remove numbers like (1), (2), (10)
    text = re.sub(r'\(\d+\)', '', text)

    # Remove extra spaces
    text = re.sub(r'\s+', ' ', text)

    return text.strip()


def extract_qna(pdf_path):
    doc = fitz.open(pdf_path)

    qna_list = []
    current_question = ""
    current_answer = ""
    capturing_question = False

    for page in doc:
        blocks = page.get_text("dict")["blocks"]

        for block in blocks:
            if "lines" not in block:
                continue

            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    font = span["font"].lower()

                    if not text:
                        continue

                    # ❌ Ignore Objective lines completely
                    if "objective" in text.lower():
                        continue

                    # ✅ Start question ONLY if bold AND contains "comments"
                    if "bold" in font and "comments" in text.lower():
                        # Save previous Q&A
                        if current_question:
                            qna_list.append((
                                clean_question(current_question),
                                current_answer.strip()
                            ))

                        # Start new question
                        current_question = text
                        current_answer = ""
                        capturing_question = True

                    # ✅ Continue question if still bold (multi-line question)
                    elif capturing_question and "bold" in font:
                        current_question += " " + text

                    # ✅ Otherwise → answer
                    elif capturing_question:
                        current_answer += " " + text

    # Save last Q&A
    if current_question:
        qna_list.append((
            clean_question(current_question),
            current_answer.strip()
        ))

    return qna_list


def save_to_excel(qna_list, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "QnA"

    ws["A1"] = "Question"
    ws["B1"] = "Answer"

    for i, (q, a) in enumerate(qna_list, start=2):
        ws.cell(row=i, column=1, value=q)
        ws.cell(row=i, column=2, value=a)

    wb.save(output_file)


# ===== RUN =====
pdf_path = "Introduction.pdf"
output_excel = "output.xlsx"

qna = extract_qna(pdf_path)
save_to_excel(qna, output_excel)

print("✅ Done! Clean Excel created.")