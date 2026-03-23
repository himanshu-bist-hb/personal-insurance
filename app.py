import fitz  # PyMuPDF
from openpyxl import Workbook

def extract_qna(pdf_path):
    doc = fitz.open(pdf_path)

    qna_list = []
    current_question = None
    current_answer = ""

    for page in doc:
        blocks = page.get_text("dict")["blocks"]

        for block in blocks:
            if "lines" not in block:
                continue

            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    font = span["font"].lower()

                    if not text:
                        continue

                    # Detect bold text (QUESTION)
                    if "bold" in font:
                        # Save previous Q&A
                        if current_question:
                            qna_list.append((current_question, current_answer.strip()))

                        # Start new question
                        current_question = text
                        current_answer = ""

                    else:
                        # Append to answer
                        if current_question:
                            current_answer += " " + text

    # Append last Q&A
    if current_question:
        qna_list.append((current_question, current_answer.strip()))

    return qna_list


def save_to_excel(qna_list, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "QnA"

    # Headers
    ws["A1"] = "Question"
    ws["B1"] = "Answer"

    for i, (q, a) in enumerate(qna_list, start=2):
        ws.cell(row=i, column=1, value=q)
        ws.cell(row=i, column=2, value=a)

    wb.save(output_file)


# ====== RUN ======
pdf_path = "Introduction.pdf"        # <-- change this
output_excel = "output.xlsx"  # <-- output file

qna = extract_qna(pdf_path)
save_to_excel(qna, output_excel)

print("Done! Data saved to Excel.")